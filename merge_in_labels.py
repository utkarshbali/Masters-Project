from openpyxl import load_workbook

wb1 = load_workbook('utkarsh-wordgraph-labels.xlsx')
# wb1.get_sheet_names()
ws1 = wb1['Sheet 1']

label = dict()

for row in ws1.rows[3:]:
	label[row[1].value] = row[2].value

wb2 = load_workbook('flattenedCompareResultsMCMC.xlsx')
ws2 = wb2['Results']

ws2.rows[0][8].value = 'Label'

for row in ws2.rows[1:]:
	word = row[0].value
	if word in label:
		#sprint "found " + word
		row[8].value = label[word]

wb2.save(filename = 'flattenedlabeled.xlsx')


